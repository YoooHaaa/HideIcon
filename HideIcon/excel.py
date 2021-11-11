# !/usr/bin/env python3
# -*-coding:utf-8 -*-

"""
# File       : excel.py
# Time       ：2021/11/5 16:31
# Author     ：Yooha
"""

import openpyxl
from error import ErrorUtil
from globalutil import Global

class ExcelUtil(object):

    def __init__(self):
        pass

    @classmethod
    def init_xlsx(cls):
        try:
            excel_sdk = openpyxl.load_workbook("out.xlsx")
            sheel_sdk = excel_sdk["hash"]
            sheel_sdk.cell(row = 1, column = 1, value = '序号')
            sheel_sdk.cell(row = 1, column = 2, value = 'hash')
            sheel_sdk.cell(row = 1, column = 3, value = 'label')
            sheel_sdk.cell(row = 1, column = 4, value = 'icon')
            excel_sdk.save("out.xlsx")
        except Exception as err:
            ErrorUtil.error('ExcelUtil', 'init_xlsx', str(err))
            return False
        return True


    @classmethod
    def write(cls, info):
        Global.global_mutex.acquire()
        try:
            excel_sdk = openpyxl.load_workbook("out.xlsx")
            sheel_sdk = excel_sdk["hash"]
            rows = sheel_sdk.max_row + 1
            sheel_sdk.cell(row = rows, column = 1, value = rows - 1)
            sheel_sdk.cell(row = rows, column = 2, value = info['hash'])
            sheel_sdk.cell(row = rows, column = 3, value = info['label'])
            sheel_sdk.cell(row = rows, column = 4, value = info['icon'])
            excel_sdk.save("out.xlsx")
        except Exception as err:
            excel_sdk.save("out.xlsx")
            Global.global_mutex.release()
            ErrorUtil.error('ExcelUtil', 'write', str(err))
            return False
        Global.global_mutex.release()
        return True
